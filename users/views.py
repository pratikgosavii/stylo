from email import message
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from users.serializer import UserProfileSerializer


from .forms import *


# def login_page(request):
#     forms = LoginForm()
#     if request.method == 'POST':
#         forms = LoginForm(request.POST)
#         if forms.is_valid():
#             username = forms.cleaned_data['username']
#             password = forms.cleaned_data['password']
#             print(username)
#             print(password)
#             user = authenticate(username=username, password=password)
#             if user:
#                 login(request, user)

#                 if user.is_vendor:
#                     print('---------------------------------')
#                     print('---------------------------------')
#                     print('---------------------------------')
#                 return redirect('dashboard')
#             else:
#                 messages.error(request, 'wrong username password')
#     context = {'form': forms}
#     return render(request, 'adminLogin.html', context)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from .models import User, OTP, DeviceToken
from .otp_utils import create_and_send_otp, verify_otp, normalize_mobile
from vendor.models import vendor_store


class SendOTPView(APIView):
    """Send OTP to mobile number"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        
        if not mobile:
            return Response({"error": "Mobile number is required"}, status=400)
        
        # Normalize mobile (same format used for storage and verify)
        mobile = normalize_mobile(mobile)
        if len(mobile) != 10:
            return Response({"error": "Invalid mobile number. Use 10 digits or 12 with 91."}, status=400)
        
        # Create and send OTP
        otp_obj, success, message = create_and_send_otp(mobile)
        
        if success:
            return Response({
                "message": "OTP sent successfully",
                "mobile": mobile
            }, status=200)
        else:
            return Response({"error": message}, status=400)


class VerifyOTPView(APIView):
    """Verify OTP and return JWT tokens"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        otp_code = request.data.get("otp")
        
        if not mobile or not otp_code:
            return Response({"error": "Mobile number and OTP are required"}, status=400)
        
        # Normalize mobile (same as send-otp)
        mobile = normalize_mobile(mobile)
        if len(mobile) != 10:
            return Response({"error": "Invalid mobile number."}, status=400)
        
        # Verify OTP
        otp_obj, is_valid, message = verify_otp(mobile, otp_code)
        
        if not is_valid:
            return Response({"error": message}, status=400)
        
        # Get or create user
        user = User.objects.filter(mobile=mobile).first()
        created = False
        
        if not user:
            user = User.objects.create(
                mobile=mobile,
                is_active=True
            )
            created = True
        
        # Generate JWT tokens
        refresh = RefreshToken.for_user(user)
        
        return Response({
            "access": str(refresh.access_token),
            "refresh": str(refresh),
            "user": {
                "id": user.id,
                "mobile": user.mobile,
                "created": created
            }
        }, status=200)


class SignupView(APIView):
    """Signup user after OTP verification"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        otp_code = request.data.get("otp")
        user_type = request.data.get("user_type")
        city = request.data.get("city")
        area = request.data.get("area")
        name = request.data.get("name")
        email = request.data.get("email")

        if not mobile or not otp_code or not user_type:
            return Response({"error": "Mobile, OTP, and user_type are required"}, status=400)

        # Normalize mobile (same as send-otp)
        mobile = normalize_mobile(mobile)
        if len(mobile) != 10:
            return Response({"error": "Invalid mobile number."}, status=400)
        
        # Verify OTP first
        otp_obj, is_valid, message = verify_otp(mobile, otp_code)
        
        if not is_valid:
            return Response({"error": message}, status=400)

        try:
            if user_type not in ("vendor", "customer"):
                return Response({"error": "Invalid user_type. Use 'vendor' or 'customer'."}, status=400)

            is_vendor = user_type == "vendor"
            is_customer = user_type == "customer"
            user = User.objects.filter(mobile=mobile).first()
            created = False

            if user:
                existing_role = "vendor" if user.is_vendor else ("customer" if user.is_customer else "user")
                if existing_role != user_type:
                    return Response({
                        "error": f"This number is already registered as a {existing_role}. Cannot register again as {user_type}."
                    }, status=400)

            else:
                # Ensure email is unique
                if email and User.objects.filter(email=email).exists():
                    return Response({"error": "This email is already in use."}, status=400)

                user = User.objects.create(
                    mobile=mobile,
                    first_name=name or "",
                    email=email or "",
                    is_vendor=is_vendor,
                    is_customer=is_customer,
                )
                created = True

            refresh = RefreshToken.for_user(user)
            return Response({
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "user": {
                    "id": user.id,
                    "mobile": user.mobile,
                    "email": user.email,
                    "name": user.first_name,
                    "user_type": user_type,
                    "created": created
                }
            })

        except Exception as e:
            return Response({"error": str(e)}, status=400)


class LoginAPIView(APIView):
    """Login user after OTP verification"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        otp_code = request.data.get("otp")
        user_type = request.data.get("user_type")

        if not mobile or not otp_code:
            return Response({"error": "Mobile number and OTP are required"}, status=400)

        # Clean mobile number
        mobile = ''.join(filter(str.isdigit, str(mobile)))
        
        # Verify OTP
        otp_obj, is_valid, message = verify_otp(mobile, otp_code)
        
        if not is_valid:
            return Response({"error": message}, status=400)

        try:
            user = User.objects.filter(mobile=mobile).first()
            created = False

            if user:
                if not user.is_active:
                    user.is_active = True
                    user.save()
            else:
                user_type = request.data.get("user_type")
                is_vendor = user_type == "vendor"
                is_customer = user_type == "customer"
                is_deliveryboy = user_type == "deliveryboy"
                user = User.objects.create(
                    mobile=mobile,
                    is_active=True,
                    is_vendor=is_vendor,
                    is_customer=is_customer,
                    is_deliveryboy=is_deliveryboy,
                )
                created = True
                if user_type == "vendor":
                    vendor_store.objects.create(user=user, name="My Store")
                    print("-------16------------ vendor_store created")

            # Token creation
            refresh = RefreshToken.for_user(user)
            user_details = UserProfileSerializer(user).data

            # Subscription status (for vendors)
            is_subscribed = False
            if user.is_vendor:
                is_subscribed = user.subscription_is_active

            role = "vendor" if user.is_vendor else ("customer" if user.is_customer else ("deliveryboy" if getattr(user, "is_deliveryboy", False) else "user"))
            return Response({
                "access": str(refresh.access_token),
                "refresh": str(refresh),
                "user": {
                    "id": user.id,
                    "mobile": user.mobile,
                    "role": role
                },
                "created": created,
                "is_subscribed": is_subscribed,
                "user_details": user_details
            }, status=201 if created else 200)

        except Exception as e:
            print(f"Login failed: {e}")
            return Response({"error": str(e)}, status=400)





from rest_framework.permissions import IsAuthenticated
from .permissions import *


class RegisterDeviceTokenView(APIView):
    """Register or update FCM device token for the logged-in user. Call after login."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        token = (request.data.get("token") or "").strip()
        if not token:
            return Response({"error": "token is required"}, status=status.HTTP_400_BAD_REQUEST)
        DeviceToken.objects.update_or_create(
            user=request.user,
            defaults={"token": token},
        )
        return Response({"message": "Device token registered"}, status=status.HTTP_200_OK)


class UsergetView(APIView):
    permission_classes = [IsCustomer]

    def get(self, request):
        user = request.user
        return Response({
            "name": user.first_name,
            "email": user.email,
        })

class UserUpdateView(APIView):
    permission_classes = [IsCustomer]

    def put(self, request):
        user = request.user
        name = request.data.get("name")
        email = request.data.get("email")

        updated = False

        if name:
            user.first_name = name
            updated = True

        if email:
            user.email = email
            updated = True

        if updated:
            user.save()
            return Response({"message": "Profile updated successfully."})
        else:
            return Response({"message": "No changes provided."}, status=status.HTTP_400_BAD_REQUEST)


class ResetPasswordView(APIView):
    """Reset password using OTP verification"""
    
    def post(self, request):
        mobile = request.data.get("mobile")
        otp_code = request.data.get("otp")
        new_password = request.data.get("new_password")

        if not mobile or not otp_code or not new_password:
            return Response({"error": "Mobile, OTP, and new_password are required"}, status=400)

        # Normalize mobile (same as send-otp)
        mobile = normalize_mobile(mobile)
        if len(mobile) != 10:
            return Response({"error": "Invalid mobile number."}, status=400)
        
        # Verify OTP
        otp_obj, is_valid, message = verify_otp(mobile, otp_code)
        
        if not is_valid:
            return Response({"error": message}, status=400)

        try:
            user = User.objects.filter(mobile=mobile).first()
            
            if not user:
                return Response({"error": "User not found"}, status=404)

            # Update password
            user.set_password(new_password)
            user.save()

            return Response({"message": "Password updated successfully."})

        except Exception as e:
            return Response({"error": str(e)}, status=400)
        



def  login_admin(request):

    forms = LoginForm()
    if request.method == 'POST':
        forms = LoginForm(request.POST)
        if forms.is_valid():
            mobile = forms.cleaned_data['mobile']
            password = forms.cleaned_data['password']
            print(mobile)
            print(password)
            user = authenticate(mobile=mobile, password=password)
            if user:
                login(request, user)

                if user.is_superuser:
                    print('---------------------------------')
                    print('---------------------------------')
                    print('---------------------------------')
                return redirect('dashboard')
            else:
                messages.error(request, 'wrong username password')
    context = {'form': forms}
    return render(request, 'adminLogin.html', context)


# def resgister_page(request):

#     forms = registerForm()
#     if request.method == 'POST':
#         forms = registerForm(request.POST)
#         if forms.is_valid():
#             forms.save()
#             username = forms.cleaned_data['username']
#             password = forms.cleaned_data['password1']
#             user = authenticate(username=username, password=password)
#             if user:
                
#                 messages.error(request, 'user already exsist')
#                 return redirect('dashboard')
#             else:
#                 return redirect('resgister')
#         else:
#             print(forms.errors)
#     else:
#         print(forms.as_p)

#         context = {'form': forms}

#         return render(request, 'users/resgister.html', context)


def logout_page(request):
    logout(request)
    return redirect('login_admin')

def user_list(request):
    data = User.objects.filter(is_vendor=True)
    return render(request, 'user_list.html', {'data': data})


def customer_list(request):
    """
    List all customers (users with is_customer=True)
    """
    data = User.objects.filter(is_customer=True).order_by('-date_joined')
    count = data.count()
    
    return render(request, 'customer_list.html', {'data': data, 'count': count})


@login_required(login_url='login_admin')
def export_customer_list_excel(request):
    """Export customer list to Excel"""
    data = User.objects.filter(is_customer=True).order_by('-date_joined')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    headers = ['#', 'User Email', 'First Name', 'Last Name', 'Mobile', 'Gender', 'Date of Birth', 'Address', 'Date Joined']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for idx, user in enumerate(data, start=1):
        ws.append([
            idx,
            user.email or "-",
            user.first_name or "-",
            user.last_name or "-",
            user.mobile or "-",
            user.gender or "-",
            user.dob.strftime("%d-%m-%Y") if user.dob else "-",
            user.address or "-",
            user.date_joined.strftime("%d-%m-%Y %H:%M") if user.date_joined else "-"
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required(login_url='login_admin')
def view_vendor_details(request, vendor_id):
    """
    View all details of a vendor (user with is_vendor=True)
    """
    user = get_object_or_404(User, id=vendor_id, is_vendor=True)
    context = {'user': user}
    return render(request, 'view_vendor.html', context)


def vendor_list(request):
    data = User.objects.filter(is_vendor=True).order_by('-date_joined')
    count = data.count()
    return render(request, 'vendor_list.html', {'data': data, 'count': count})


@login_required(login_url='login_admin')
def export_vendor_list_excel(request):
    """Export vendor list to Excel with subscription valid till date"""
    data = User.objects.filter(is_vendor=True).order_by('-date_joined')

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendors"
    headers = ['#', 'User Email', 'Name', 'Mobile', 'Gender', 'Address', 'Subscription Valid From', 'Subscription Valid Till']
    ws.append(headers)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, user in enumerate(data, start=1):
        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or "-"
        address = user.address or "-"
        ws.append([
            idx,
            user.email or "-",
            full_name,
            user.mobile or "-",
            user.gender or "-",
            address,
            user.subscription_valid_from.strftime("%d-%m-%Y") if user.subscription_valid_from else "-",
            user.subscription_valid_to.strftime("%d-%m-%Y") if user.subscription_valid_to else "-"
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="vendors_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login_admin')
def update_user_subscription(request, user_id):
    """
    View and update user subscription details (for vendors)
    """
    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Handle user account activate/deactivate
        if action == 'activate_user':
            user.is_active = True
            user.save()
            messages.success(request, 'User account activated successfully!')
            return redirect('update_user_subscription', user_id=user_id)
        
        elif action == 'deactivate_user':
            user.is_active = False
            user.save()
            messages.success(request, 'User account deactivated successfully!')
            return redirect('update_user_subscription', user_id=user_id)
        
        # Regular update of subscription fields
        subscription_valid_from = request.POST.get('subscription_valid_from')
        subscription_valid_to = request.POST.get('subscription_valid_to')
        subscription_received_amount = request.POST.get('subscription_received_amount')
        
        if subscription_valid_from:
            user.subscription_valid_from = subscription_valid_from
        else:
            user.subscription_valid_from = None
            
        if subscription_valid_to:
            user.subscription_valid_to = subscription_valid_to
        else:
            user.subscription_valid_to = None
        
        if subscription_received_amount:
            try:
                user.subscription_received_amount = float(subscription_received_amount)
            except (ValueError, TypeError):
                pass
        else:
            user.subscription_received_amount = 0.00
            
        user.save()
        
        messages.success(request, 'Subscription details updated successfully!')
        return redirect('update_user_subscription', user_id=user_id)

    context = {'user': user}
    return render(request, 'update_user_subscription.html', context)


@login_required(login_url='login_admin')
def subscription_payment_history(request, user_id):
    """
    View payment history for user subscription (vendors)
    """
    user = get_object_or_404(User, id=user_id)
    payment_history = []
    total_received = 0

    context = {
        'user': user,
        'payment_history': payment_history,
        'total_received': total_received,
    }
    return render(request, 'subscription_payment_history.html', context)




from rest_framework import viewsets, mixins
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.decorators import action


class UserProfileViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def _get_profile_completed(self, user):
        """For vendors: True if vendor_store exists and has key fields filled. Else True for non-vendors."""
        if not getattr(user, "is_vendor", False):
            return True
        try:
            store = vendor_store.objects.get(user=user)
        except vendor_store.DoesNotExist:
            return False
        # Check key vendor_store fields are present
        key_fields = [
            "name", "city", "pincode", "business_type",
            "house_building_no", "locality_street", "state",
            "vendor_house_no", "vendor_locality_street", "vendor_pincode", "vendor_state", "vendor_city",
            "latitude", "longitude",
        ]
        for field in key_fields:
            val = getattr(store, field, None)
            if val is None or (isinstance(val, str) and not val.strip()):
                return False
        return True

    @action(detail=False, methods=['get', 'put'], url_path='me')
    def me(self, request):
        user = request.user

        if request.method == 'GET':
            serializer = UserProfileSerializer(user)
            data = dict(serializer.data)
            data["profile_completed"] = self._get_profile_completed(user)
            return Response(data)

        elif request.method == 'PUT':
            serializer = UserProfileSerializer(user, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                data = dict(serializer.data)
                data["profile_completed"] = self._get_profile_completed(user)
                return Response(data)

            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        





from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction, IntegrityError


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_user(request):
    """Delete user account"""
    user = request.user

    try:
        with transaction.atomic():
            # Delete user from Django
            user.delete()

        return Response(
            {"detail": "Your account has been deleted successfully."},
            status=status.HTTP_204_NO_CONTENT
        )

    except IntegrityError as e:
        return Response(
            {"detail": f"Account deletion failed: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST
        )
